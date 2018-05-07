#coding=utf-8


'''
需求：
    1.提前一天通知上班时间
    2.当天上班前一小时通知

思路:
    1.检查本地时间，定时发送通知
        1.1 设置一个死循环一直检测本地时间
        1.2 设置触发时间
            提前一天通知需要在晚上20:00 通知
            当天上版通知需要提前1小时通知
    2.获取excel中的数据
        提前一天通知需要，遍历第二天所有的人
        当天通知需要，检查指定班次通知。
    

'''

import requests
import time
from openpyxl import load_workbook


def getToken():

    username = 'zhangyinkai'
    domain_name = 'hwcloudsom1'
    password = 'zYK#1410'
    project_name = "cn-north-1"
    IAM_URL = 'https://iam.cn-north-1.myhuaweicloud.com/v3/auth/tokens'
    headers = {'Content-Type':'application/json;charset=utf8'}

    body = {
      "auth": {
        "identity": {
          "methods": [
            "password"
          ],
          "password": {
            "user": {
              "name": username,
              "password": password,
              "domain": {
                "name": domain_name
              }
            }
          }
        },
        "scope": {
          "project": {
            "name": project_name

          }
        }
      }
    }

    r = requests.post(url=IAM_URL,json=body,headers=headers)
    # print r.status_code

    # print r.headers
    # print r.headers['X-Subject-Token']
    return r.headers['X-Subject-Token']

def sendMessages(phones,messages,sign_id = "e0cde4e0b08e4ce590344737e852abe5"):
    '''
    发送短信通知
    :param phones: 需要通知的列表
    :param messages: 通知内容
    :param sign_id: 短信签名
    :return:
    '''

    # phones = ["13323952520","13403622324"]
    # messages = "上班提醒："
    # sign_id = "e0cde4e0b08e4ce590344737e852abe5"
    SDM_URL = ' https://smn.cn-north-1.myhuaweicloud.com/v2/52fb7d7429d04068ae8ff9632106e701/notifications/sms'

    tokens = getToken()
    headers = {
        'Content-Type': 'application/json;charset=utf8',
        'X-Auth-Token': tokens
    }
    body =  {
        "endpoints": phones,
        "message": messages,
        "sign_id": sign_id
    }

    r = requests.post(url=SDM_URL,json=body,headers=headers)
    # print r.status_code

def getEmployeePhone(filename,phoneSheet = 'phone'):
    '''
    获取姓名和手机号的对应关系
    :param filename:  保存姓名和手机号的对应关系的文件名称
    :param phoneSheet: 保存姓名和手机号的对应关系的文件的工作表的名称
    :return: 返回姓名和手机号的对应关系字典
    '''


    wb = load_workbook(filename)
    ws = wb[phoneSheet]

    phoneList = {}

    # 设置人数的总量70，必须要保证人数小于70，文件的第一列是姓名，第二列是手机号
    for i in range(1,70):
        phoneList[ws.cell(row=i, column=1).value] = ws.cell(row=i, column=2).value

    # for i in phoneList:
    #     print i,phoneList[i]

    return phoneList


def getstafflist(filename,today,tomorrow):
    '''
    查询当天的上班情况、明天的上班情况，返回上班情况的字典
    :param filename:  值班的excel表
    :param today: 今天日期 使用unicode
    :param tomorrow: 明天的日期 使用unicode
    :return:  返回列表， 今天和明天对应人员的上班情况
    '''

    '''
    

    编码格式有问题，需要修改
    :return:
    '''
    # filename = 't2.xlsx'
    # today = u'5日'
    # tomorrow = u'6日'
    today_list = {}
    tomorrow_list = {}
    nameCol = 0


    wb = load_workbook(filename)
    ws = wb['work']

    for i in range(1,40):
        cellValue = ws.cell(row=1, column=i).value
        # print cellValue
        # cellValue = cellValue.decode('utf-8')
        # print cellValue
        if cellValue == None:
            continue
        if cellValue == u'姓名':
            nameCol = i
            # print "#######xingming######"
            # print nameCol

        elif cellValue == today:
            todayCol = i

            # 查询60列，排班总人数少于60人
            for num in range(1,60):
                today_list[ws.cell(row=num,column=nameCol).value] =  ws.cell(row=num,column=i).value

        elif cellValue == tomorrow:
            # 查询60列，排班总人数少于60人
            for num in range(1, 60):
                tomorrow_list[ws.cell(row=num, column=nameCol).value] = ws.cell(row=num, column=i).value

    # print today_list
    # print "########getstafflist########tomorrow_list"
    # print ws['B4'].value
    # print  tomorrow_list
    return today_list,tomorrow_list


def notice_all(fileName):


    # fileName = 't1.xlsx'

    now_time = time.strftime("%Y-%m-%d ", time.localtime())
    now_time_hour = time.strftime("%H", time.localtime())
    now_time_day = time.strftime("%d", time.localtime())


    # 类型转化
    today = str(int(now_time_day)) + u'日'
    tomorrow = str(int(now_time_day) + 1) + u'日'

    todaylist,tomorrowlist = getstafflist(fileName,today,tomorrow)


    bai1 = []
    bai2 = []
    xia = []
    xiaE = []
    yeban = []
    xiuxi = []

    # print "#########noticeall######tomorrowlist"
    # print  tomorrowlist
    for shifts in tomorrowlist:
        if tomorrowlist[shifts] == u'白1':
            bai1.append(shifts)
            pass
        elif tomorrowlist[shifts] == u'白2':
            bai2.append(shifts)
            pass
        elif tomorrowlist[shifts] == u'下':
            xia.append(shifts)
            pass
        elif tomorrowlist[shifts] == u'下E':
            xiaE.append(shifts)
            pass
        elif tomorrowlist[shifts] == u'夜':
            yeban.append(shifts)
            pass
        elif tomorrowlist[shifts] == u'修':
            xiuxi.append(shifts)
            pass

    # print bai1,bai2
    # print yeban
    # print xia,xiaE
    # print xiuxi

    # 需要输入保存姓名和手机号的对应关系的文件名、工作簿的名称（默认是phone），
    phoneList = getEmployeePhone(fileName)


    bai1PhoneList = []
    bai2PhoneList = []
    xiaPhoneList = []
    xiaEPhoneList = []
    yebanPhoneList = []
    xiuxiPhoneList = []

    print bai1
    for i in bai1:
        bai1PhoneList.append(phoneList[i])

    for i in bai2:
        bai2PhoneList.append(phoneList[i])

    for i in xia:
         xiaPhoneList.append(phoneList[i])

    for i in xiaE:
        xiaEPhoneList.append(phoneList[i])

    for i in yeban:
        yebanPhoneList.append(phoneList[i])

    for i in xiuxi:
        xiuxiPhoneList.append(phoneList[i])


    bai1Message = '上班提醒：今天' + now_time + '明天您白1班（8:30-18:30）'
    bai2Message = '上班提醒：今天' + now_time + '明天您白2班（9:30-19:30）'
    xiaMessage = '上班提醒：今天' + now_time + '明天您下午班（14:00-23:00）'
    xiaEMessage = '上班提醒：今天' + now_time + '明天您下午班（17:00-22:00）'
    yebanMessage = '上班提醒：今天' + now_time + '明天您下午班（21:00-次9:00）'

    sendMessages(bai1PhoneList,bai1Message)
    sendMessages(bai2PhoneList,bai2Message)
    sendMessages(xiaPhoneList,xiaMessage)
    sendMessages(xiaEPhoneList,xiaEMessage)
    sendMessages(yebanPhoneList,yebanMessage)

    # print bai1PhoneList,bai1Message

if __name__ == '__main__':

    # getstafflist('t1.xlsx',u'6日',u'7日')
    # getEmployeePhone()
    # sendMessages(['13403622324'],'send test')
    fileName = 't2.xls'

    while True:
    	now_time = time.strftime("%H:%M", time.localtime())
    	# print now_time
    	if str(now_time) == '23:48':
            notice_all(fileName)
            time.sleep(60)
