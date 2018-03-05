#coding=utf-8
"""
需求：将excel表中的用户名注入到统一身份认证中

思路：
	1.使用python的模块openpyxl将excel表中的数据读取，定义方法：get_users_xls
	2.调用API使用excel表中的数据创建用户会返回用户ID
	3.获取IAM中组和组的ID
	4.通过组的名称查询对应的ID
	5.增加用户到指定的组中



"""

import sys
import requests
from openpyxl import load_workbook

reload(sys)

sys.setdefaultencoding('utf-8')

def get_Token(username, password, region):
    """
    获取token
    :param username:
    :param password:
    :param region:
    :return:
    """
    url = 'https://iam.' + region + '.myhwclouds.com/v3/auth/tokens'

    headers = {'Content-Type': 'application/json;charset=utf8'}

    data = {
        "auth": {
            "identity": {
                "methods": [
                    "password"
                ],
                "password": {
                    "user": {
                        "name": username,
                        "password": password,
                        "domain": {
                            "name": username
                        }
                    }
                }
            },
            "scope": {
                "domain": {
                    "name": username
                }
            }
        }
    }

    r = requests.post(url, json=data, headers=headers)
    # print r.headers
    return r.headers['X-Subject-Token']

def list_iam_users():
    """
    显示统一身份认证中所有的用户信息
    :return:用户和用户ID组成的字典
    """

    url = host + '/v3/users'
    r = requests.get(url,headers=headers)
    comp = r.json()
    users_dic = {}
    for name in comp.get('users'):
        users_dic[name.get('name')] = name.get('id')

    return users_dic

def add_iam_user(iam_user,iam_user_pw='Jfz!955988',iam_user_email=''):
    """
    在统一身份认证中增加用户
    :param iam_user: 增加的用户名,必须要大于五个字符串
    :param iam_user_pw: 用户名对应的密码
    :param iam_user_email: 用户名对应的邮箱
    :return: 用户ID
    """

    try:
        url = host + '/v3/users'
        data = {
            "user": {
                "name":iam_user,
                "password":iam_user_pw,
                "email":iam_user_email
            }
            }
        r = requests.post(url,json=data, headers=headers)
        # print iam_user
        # print r.json()

        if r.json().get('error'):
            raise ValueError,r.json().get('error').get('message')

    except ValueError,reason:
        print "错误用户名：" + iam_user
        print '错入提示：', reason
        exit(1)
    else:
        return r.json().get('user').get('id')

def get_users_xls(filename):
    """
    华为邮件获取用户名的拼音，返回用户名和邮箱组成的字典
    :param filename:Execl表的名称
    :return: 用户名和邮箱的字典
    """
    wb = load_workbook(filename)
    ws = wb.active

    # 表中的第一列是用户名的拼音
    ROW = 2
    COL = 3
    USERDIC = {}
    while True:
        if ws.cell(row=ROW,column=COL).value == None:
            break

        else:
            email = ws.cell(row=ROW, column=COL).value
            user_add = email.split('@')[0]
            USERDIC[user_add] = email
            ROW += 1

    return USERDIC


def list_iam_groups():
    """
    获取统一身份认证中的组以及对应的组ID
    :return: 组以及对应的组ID组成的字典
    """

    url = 'https://iam.cn-north-1.myhwclouds.com/v3/groups'
    # url = host + '/v3/groups'
    r = requests.get(url, headers=headers)
    comp = r.json()
    group_dic = {}
    for group_name in comp.get('groups'):
        group_dic[group_name.get('name')] = group_name.get('id')

    return group_dic


def add_usertogroup(userID,groupID):
    """
    将用户添加到指定的组中
    :param userID: IAM用户ID
    :param groupID: IAM组ID
    :return:
    """
    url = 'https://iam.cn-north-1.myhwclouds.com/v3/groups/' + groupID + '/users/'+ userID
    # url = host + '/v3/groups/' + groupID + '/users/'+ userID
    r = requests.put(url, headers=headers)
    return 0

def import_xlsx_to_iamgroup(fileName,group_name):
    """
    输入用户信息的excel表名和需要加入到的组，该文件中第三列必须是用户的邮箱，
    :param fileName:用户信息的excel表名
    :param group_name:组名
    :return:
    """
    user_email = get_users_xls(fileName)
    for iam_user_name in user_email:
        userID = add_iam_user(iam_user_name,user_email[iam_user_name],user_email[iam_user_name])
        groupDic = list_iam_groups()
        if 0 == add_usertogroup(userID,groupDic[group_name]):
            print iam_user_name + "注入成功"


if __name__ == '__main__':
    # 账号和密码(修改)
    username = "********"
    password = "********"

    # 项目ID和查询的区域（修改）
    area = ('cn-north-1', 'cn-south-1', 'cn-east-2')
    project_set = {'cn-north-1': '52fb7d7429d04068ae8ff9632106e701', 'cn-south-1': 'ee766306d2ad4387a56a0e31172185b5',
                   'cn-east-2': '21275af600cd4305840e5d10a500ee87'}

    # 选择对应的区域
    region = area[1]
    project_id = project_set[region]

    host = 'https://iam.' + region + '.myhuaweicloud.com'

    # 获取token
    tokens = get_Token(username, password, region)
    headers = {
        'Content-Type': 'application/json;charset=utf8',
        'X-Auth-Token': tokens
    }


    # 输入保存用户名的excel表名，和需要加入的组名(修改)
    FILENAME = 'test.xlsx'
    GROUPNAME = 'admin'
    import_xlsx_to_iamgroup(FILENAME,GROUPNAME)




