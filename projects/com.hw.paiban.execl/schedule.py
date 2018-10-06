#coding=utf-8
#
#

'''

需求： 
    1.根据组内排班情况输出根据固定模板的execl表
    
实现步骤：
    1.获取排班情况。
    2.将组ID、人员、设置成固定变量。
    3.通过获取的员工信息创建实例
    4.将员工的信息设置，最后输出到一个指定的文件output

# file_name: 调用的execl表的文件名
# sheet_name: 调用sheet的名称
'''

from group import Group
from employee import Employee
from openpyxl import load_workbook
from datetime import date
import sys
reload(sys)
sys.setdefaultencoding('utf-8')



def getstafflist(filename,today,sheetname='work',output=2):
    '''
     
    查询当天的上班情况，返回以名字为关键字，指定字段对应的值为值

    :param filename: 字符串，查询的execl表 
    :param today: 字符串，查询的指定列的首字段
    :param sheetname: 字符串，查询指定的sheet的名称
    :param output: 整形，1,2输出几个列表，
    :return: 字典，输出以姓名列字段未关键字，指定字段为值的字典
    '''


    '''

    编码格式有问题，需要修改
    :return:
    '''
    # filename = 't2.xlsx'
    # today = u'5日'
    # tomorrow = u'6日'
    today_list = {}
    today_work = {}
    nameCol = 0


    wb = load_workbook(filename)
    ws = wb[sheetname]

    # 按照行取值，i表示的是列号
    for i in range(1,62):
        cellValue = ws.cell(row=1, column=i).value
        # print cellValue
        # cellValue = cellValue.decode('utf-8')
        # print cellValue
        if cellValue == None:
            continue
        if cellValue == u'姓名':
            nameCol = i

        elif cellValue == today:
            todayCol = i

            # 查询60行，排班总人数少于60人
            for num in range(1,60):

                if ws.cell(row=num,column=nameCol).value == None or ws.cell(row=num,column=nameCol).value == u'姓名':
                    continue
                else:
                    today_list[ws.cell(row=num,column=nameCol).value] = ws.cell(row=num,column=i).value
                    today_work[ws.cell(row=num,column=nameCol).value] = ws.cell(row=num,column=i+1).value


    if output == 2:
        return today_list,today_work
    elif output == 1:
        return today_list
    else:
        raise ValueError,"The value of output is error"

def setGroup(file_Name='',sheet_name=''):
    """
    设置组的属性,每一行遍历一次，将数据存储到类中。
    :param fileName: 字符串，有写组名和组ID已经组的不同班的ID
    :param sheet_name: 字符串，不同组的sheet名
    :return: 两个字典，一个是当天的工作、当天的上班情况。
    """

    wb1 = load_workbook(file_Name)
    work = wb1[sheet_name]

    # 对组类赋值
    group1 = Group()
    dir1 = {}



    for row in range(1,40):
        row_list = []
        for col in range(1,6):

            cellValue = work.cell(row=row, column=col).value
            # print cellValue
            row_list.append(cellValue)




        if row_list[4] is not None:
           group1.setGroupID(row_list[4])

        dir1[row_list[1]] = row_list[0]

        group1.setShift(dir1)
        if row_list[3] is not None:
            group1.setGroupName(row_list[3])
        # print row_list[3]
    return group1


def getPersons(group,staff):
    '''
    输入员工的对象列表和需要查询的班次，输出对应的工号列表
    :param group: 对象列表，员工对象列表
    :param staff: 字符串，查询的班次
    :return:  字符串，查询指定班次的列表，
    
    例如，staff 是白1  需要查询是白1 的所有员工的工号列表。
    '''

    list = ''

    for person in group:
        # print person.staff,staff
        if person.staff == staff:
            # print type(person.today_work)
            # print "$$$$$$"
            # print type(person.staff)
            # print "#####"
            if staff.count('夜'):
                list = list + person.ID + ','
            else:
                if person.today_work == None:
                    continue
                elif "400" in person.today_work:
                    list = list + person.ID + ','

    return list[:-1]



def getStaffID(group,banci):
    '''
    输入类的实例，获取组的
    :param group: 组类，
    :param banci: 字符串，获取
    :return: 字符串，班次ID
    '''

    # 查询班次ID
    for staff in group.shift:
        # print staff,banci
        if staff.count(banci):
            # print group.shift[staff]
            return group.shift[staff].rstrip()


if __name__ == '__main__':

    # 测试一： 获取人名、工作排班
    file1 = 't1.xlsx'

    # 不同组选择不同的sheetname，存放不同的组ID排班ID
    groupName = 'groupC'


    # 设置输出文档的开始的行
    out_file_row = 2

    # 设置查询的时间
    year = 2018
    mouth = 9

    # 设置输出的文件名称
    outfile_name = '3zu.xlsx'

    # 打开最终的导出数据的表格
    out_file = 'output.xlsx'
    wb = load_workbook(out_file)
    ws = wb['数据验证']



    for day in range(1,32):

        try:
            tmp_date = date(year, mouth, day).isoformat()

        except ValueError, reson:
            if reson.message == "day is out of range for month":
                break
        else:
            # print "##"+tmp_date
            find_day = str(day) + '日'
            day_in_week = date(year,mouth,day).isoweekday()


            # 获取组的信息
            g3 = setGroup(file1, groupName)


            # 获取员工和工号的对应信息
            person_dir = getstafflist(file1,'OSM工号','employee',1)

            people_list = []
            for person in person_dir:
                people_list.append(Employee(person,person_dir[person],g3.groupID))


            # work_tab 上班情况
            # working_arr 工作安排
            work_tab, working_arr = getstafflist(file1, find_day)

            # 设置每个人的工作安排和班次
            for name1 in work_tab:
                for name2 in working_arr:
                    if name2 == name1:
                        # print name1, working_arr[name1], work_tab[name1]

                        for p1 in people_list:
                            name = p1.Name
                            name = name.rstrip()

                            if name1 == name:
                                # print name1,p1.Name
                                p1.setWork(working_arr[name2])
                                p1.setStaff(work_tab[name1])



            # 查询白2的所有人
            for banci in ('白1','白2','下','下E','夜'):

                list1 = getPersons(people_list, banci)
                if len(list1) >= 1:
                    if day_in_week == 6 or day_in_week == 7:
                        if banci == '白1':
                            banci = "节假日白班"

                    # 获取指定班次的ID
                    staffID = getStaffID(g3, banci)

                    # print g3.groupID, tmp_date, staffID, list1
                    ws.cell(row=out_file_row, column=1).value = g3.groupID
                    ws.cell(row=out_file_row, column=2).value = tmp_date
                    ws.cell(row=out_file_row, column=3).value = staffID
                    ws.cell(row=out_file_row, column=4).value = list1
                    out_file_row += 1

    wb.save(outfile_name)















